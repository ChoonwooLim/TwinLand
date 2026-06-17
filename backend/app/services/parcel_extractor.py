"""업로드 파일 → 필지(동/리·지번) 목록 추출.

지목·면적 등 권위 속성은 추출하지 않는다(프론트의 VWorld 자동조회가 담당).
LLM 호출은 OpenClaw LAN 게이트웨이만 사용 — API 키 금지.
"""
from __future__ import annotations

import csv
import io
import json
import re

# 전각 숫자/하이픈 → 반각
_FULLWIDTH = {ord("０") + i: chr(ord("0") + i) for i in range(10)}
_FULLWIDTH[ord("－")] = "-"
_FULLWIDTH[ord("—")] = "-"

# 지번 형태: (산)? 숫자 (-숫자)?
LOT_RE = re.compile(r"산?\s?\d+(?:\s?-\s?\d+)?")


def normalize_lot(s: str) -> str:
    s = (s or "").translate(_FULLWIDTH).strip()
    s = re.sub(r"\s+", "", s)              # 내부 공백 제거 ("산 31" → "산31")
    return s


def lenient_json(text: str):
    """LLM 응답에서 JSON 배열을 관대하게 파싱. 실패 시 None."""
    if not text:
        return None
    t = text.strip()
    # 코드펜스 제거
    t = re.sub(r"^```(?:json)?\s*", "", t)
    t = re.sub(r"\s*```$", "", t).strip()
    try:
        return json.loads(t)
    except Exception:
        pass
    # 첫 '[' ~ 마지막 ']' 슬라이스
    i, j = t.find("["), t.rfind("]")
    if i != -1 and j != -1 and j > i:
        try:
            return json.loads(t[i : j + 1])
        except Exception:
            return None
    return None


class ExtractError(ValueError):
    pass


# 헤더 후보 (소문자/공백제거 비교)
_LOT_HEADERS = {"지번", "지번주소", "번지", "지번번호", "lot", "jibun", "본번부번"}
_LOC_HEADERS = {"동리", "동/리", "리", "소재지", "법정동", "location", "동", "읍면동"}
_CAT_HEADERS = {"지목", "지목명", "category", "용도"}
_AREA_HEADERS = {"면적", "면적㎡", "면적(㎡)", "면적m2", "면적(m2)", "area", "area_m2", "공부면적"}
_OWNER_HEADERS = {"소유자", "소유주", "소유자명", "owner", "성명"}

# "도로", "임야" 등 면적 셀 앞에 붙는 지목 글자 (면적칸에 지목+면적이 합쳐진 표 대응)
_JIMOK_WORDS = {
    "전", "답", "과수원", "목장용지", "임야", "광천지", "염전", "대", "공장용지", "학교용지",
    "주차장", "주유소용지", "창고용지", "도로", "철도용지", "제방", "하천", "구거", "유지",
    "양어장", "수도용지", "공원", "체육용지", "유원지", "종교용지", "사적지", "묘지", "잡종지", "대지",
}


def _pick_col(header: list[str], candidates: set[str]) -> int:
    for i, h in enumerate(header):
        key = re.sub(r"\s+", "", (h or "")).lower()
        if key in candidates:
            return i
    return -1


def parse_area_m2(val) -> float:
    """면적 값 → ㎡ 숫자. '2,044m2'·'48㎡'·48.0 등 허용. 실패 시 0."""
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val or "")
    m = re.search(r"[\d,]+(?:\.\d+)?", s)
    if not m:
        return 0.0
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return 0.0


def _split_cat_area(cell: str) -> tuple[str, float]:
    """'도로 48m2' 처럼 지목+면적이 합쳐진 셀을 (지목, 면적) 으로 분리."""
    s = str(cell or "").strip()
    cat = ""
    for w in s.split():
        if w in _JIMOK_WORDS:
            cat = w
            break
    return cat, parse_area_m2(s)


def _rows_from_table(table: list[list[str]]) -> list[dict]:
    if not table:
        raise ExtractError("빈 표")
    header = [str(c) for c in table[0]]
    lot_i = _pick_col(header, _LOT_HEADERS)
    if lot_i == -1:
        raise ExtractError("지번 컬럼을 찾지 못함 (헤더: %s)" % header)
    loc_i = _pick_col(header, _LOC_HEADERS)
    cat_i = _pick_col(header, _CAT_HEADERS)
    area_i = _pick_col(header, _AREA_HEADERS)
    owner_i = _pick_col(header, _OWNER_HEADERS)

    def cell(row, i):
        return str(row[i] or "").strip() if 0 <= i < len(row) else ""

    out = []
    for raw in table[1:]:
        if lot_i >= len(raw):
            continue
        lot = normalize_lot(str(raw[lot_i] or ""))
        if not lot:
            continue
        category = cell(raw, cat_i)
        area_m2 = parse_area_m2(cell(raw, area_i)) if area_i != -1 else 0.0
        # 면적 칸에 '도로 48m2' 처럼 지목이 섞여 있고 별도 지목칸이 없으면 분리
        if area_i != -1 and (not category or area_m2 == 0):
            sc, sa = _split_cat_area(cell(raw, area_i))
            category = category or sc
            area_m2 = area_m2 or sa
        out.append({
            "location": cell(raw, loc_i),
            "lot": lot,
            "category": category,
            "area_m2": area_m2,
            "owner": cell(raw, owner_i),
        })
    return out


def extract_tabular(data: bytes, ext: str) -> list[dict]:
    if ext == "csv":
        text = data.decode("utf-8-sig", errors="replace")
        table = [row for row in csv.reader(io.StringIO(text))]
    elif ext in ("xlsx", "xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        table = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    else:
        raise ExtractError(f"지원하지 않는 표 형식: {ext}")
    return _rows_from_table(table)


def _looks_like_lot(lot: str) -> bool:
    """행번호·페이지번호 같은 단독 1~2자리 숫자를 지번 후보에서 제외.
    산 접두사·부번(하이픈)·3자리 이상 본번만 지번으로 인정(텍스트 PDF 오탐 완화)."""
    if lot.startswith("산") or "-" in lot:
        return True
    return len(lot) >= 3


def extract_lots_from_text(text: str) -> list[dict]:
    """텍스트(텍스트PDF 추출 결과)에서 지번 후보를 정규식으로 수집(중복 제거, 순서 유지)."""
    seen = set()
    out = []
    for m in LOT_RE.finditer(text or ""):
        lot = normalize_lot(m.group(0))
        if lot and lot not in seen and _looks_like_lot(lot):
            seen.add(lot)
            out.append({"location": "", "lot": lot})
    return out


def extract_text_pdf(data: bytes) -> tuple[list[dict], bool]:
    """텍스트 PDF → (rows, is_scanned). 텍스트가 비면 스캔으로 간주(rows=[], True)."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    chunks = []
    for page in reader.pages[:15]:
        try:
            t = page.extract_text() or ""
        except Exception:
            continue
        if t.strip():
            chunks.append(t)
    text = "\n".join(chunks).strip()
    if not text:
        return [], True
    return extract_lots_from_text(text), False


def pdf_first_page_png(data: bytes) -> bytes | None:
    """스캔 PDF 첫 페이지를 PNG 바이트로. pdf2image/poppler 부재 시 None."""
    try:
        from pdf2image import convert_from_bytes
    except ImportError:
        return None
    try:
        images = convert_from_bytes(data, first_page=1, last_page=1, dpi=200)
    except Exception:
        return None
    if not images:
        return None
    buf = io.BytesIO()
    images[0].save(buf, format="PNG")
    return buf.getvalue()
